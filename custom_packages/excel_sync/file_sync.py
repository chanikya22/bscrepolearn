"""
FILE_DOWNLOAD path: SharePoint .xlsx binary -> S3 -> cached sheet values -> Postgres.

Used when Graph /workbook/usedRange would timeout (large files, formulas,
external workbook links). Onboarding is still just a SheetSyncConfig.
"""
import logging
import os
import tempfile
from datetime import datetime, timezone, timedelta

import pandas as pd
from openpyxl import load_workbook

from .auth import get_graph_token
from .config import SheetSyncConfig
from .loader import full_refresh_load
from .sharepoint_reader import SharePointExcelReader
from utils.postgresconnector_v3 import PostgresConnector
from utils.s3utility import upload_file_to_s3

logger = logging.getLogger(__name__)

IST = timezone(timedelta(hours=5, minutes=30))
EXCEL_ERRORS = {"#N/A", "#N/A!", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA"}
PLACEHOLDERS = {"", "-", "--", "NA", "N/A", "None", "none", "null"}


def _prepend_row(first, rows):
    yield first
    yield from rows


def _resolve_sheet_name(requested: str, available: list) -> str:
    if requested in available:
        return requested
    folded = requested.casefold()
    for name in available:
        if name.casefold() == folded:
            return name
    key = "".join(ch for ch in requested.lower() if ch.isalnum())
    matches = [name for name in available if "".join(ch for ch in name.lower() if ch.isalnum()) == key]
    if len(matches) == 1:
        return matches[0]
    raise ValueError(f"Sheet '{requested}' not found. Available: {available}")


def _clean_header(header, idx: int) -> str:
    if header is None or str(header).strip() == "":
        return f"unnamed_{idx}"
    return " ".join(str(header).replace("\n", " ").split())


def _blank_to_none(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in PLACEHOLDERS or stripped.upper().replace(" ", "") in EXCEL_ERRORS:
            return None
    return value


def _excel_serial_to_datetime(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return series
    out = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    numeric = pd.to_numeric(series, errors="coerce")
    serial_mask = numeric.between(20000, 60000)
    if serial_mask.any():
        out.loc[serial_mask] = pd.to_datetime(numeric.loc[serial_mask], unit="d", origin="1899-12-30")
    remaining = out.isna() & series.notna() & ~serial_mask
    if remaining.any():
        parsed = pd.to_datetime(series.loc[remaining], errors="coerce")
        if pd.api.types.is_datetime64_any_dtype(parsed):
            parsed = parsed.mask((parsed.dt.year == 1970) & (parsed.dt.month == 1) & (parsed.dt.day == 1))
        out.loc[remaining] = parsed
    return out


def read_sheet(path: str, sheet_name: str, header_row: bool = True) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        resolved = _resolve_sheet_name(sheet_name, workbook.sheetnames)
        if resolved != sheet_name:
            logger.info("Using sheet '%s' for configured name '%s'", resolved, sheet_name)
        rows = workbook[resolved].iter_rows(values_only=True)
        first = next(rows, None)
        if first is None:
            return pd.DataFrame()
        if header_row:
            columns = [_clean_header(h, i) for i, h in enumerate(first)]
        else:
            columns = [f"col_{i}" for i in range(len(first))]
        width = len(columns)
        data = []
        data_rows = rows if header_row else _prepend_row(first, rows)
        for row in data_rows:
            values = [_blank_to_none(cell) for cell in row]
            if not any(cell is not None for cell in values):
                continue
            if len(values) < width:
                values.extend([None] * (width - len(values)))
            else:
                values = values[:width]
            data.append(values)
        df = pd.DataFrame(data, columns=columns)
        for col in df.columns:
            if "date" in str(col).lower():
                df[col] = _excel_serial_to_datetime(df[col])
        return df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    finally:
        workbook.close()


def _s3_key(prefix: str, filename: str) -> str:
    now = datetime.now(IST)
    prefix = prefix.strip("/")
    return f"{prefix}/year={now.year}/month={now.month:02d}/day={now.day:02d}/{filename}"


def run_file_sync(config: SheetSyncConfig, run_id=None) -> dict:
    filename = os.path.basename(config.file_path)
    s3_key = _s3_key(config.s3_prefix, filename)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        logger.info("Downloading %s from SharePoint", config.file_path)
        SharePointExcelReader(get_token_fn=get_graph_token).download_file(
            config.site_id, config.file_path, tmp_path
        )
        if os.path.getsize(tmp_path) == 0:
            raise RuntimeError(f"Downloaded empty file: {config.file_path}")
        upload_file_to_s3(tmp_path, config.s3_bucket, s3_key)
        df = read_sheet(tmp_path, config.sheet_name, header_row=config.header_row)
        logger.info("Read %s rows x %s cols from %s", len(df), len(df.columns), config.sheet_name)
        stats = full_refresh_load(
            PostgresConnector(), df, config.target_schema, config.target_table, run_id=run_id
        )
        stats["s3_uri"] = f"s3://{config.s3_bucket}/{s3_key}"
        return stats
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
